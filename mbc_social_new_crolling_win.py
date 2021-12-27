from sys import hash_info
import time
from selenium import webdriver
#from pyvirtualdisplay import Display
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
import chromedriver_autoinstaller
from selenium.webdriver.common.keys import Keys
import openpyxl
from datetime import datetime
from datetime import date, timedelta
#===================== 메일 발송 패키지
import os
import smtplib
from email.encoders import encode_base64
from email.header import Header
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import threading
from time import process_time, sleep
import signal


#드라이버 객체 반환
def setting_driver():
    #chromedriver_autoinstaller.install()
    #print("---chromedriver_auto install done")
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument('--headless')
    chrome_options.add_argument('window-size=1920,1400')
    #chrome_options.add_argument("--single-process")
    #chrome_options.add_argument("--disable-dev-shm-usage")
    driver = webdriver.Chrome("chromedriver",chrome_options=chrome_options)
    print("---chromedriver import")
    driver.set_window_position(0, 0)
    driver.set_window_size(1920, 1080)
    driver.implicitly_wait(10)

    return driver

#엑셀 파일 저장
def create_xls(sheet_title, file_name):
    wb = openpyxl.Workbook()
    sheet1 = wb.active
    sheet1.title = sheet_title
    new_file_name = file_name

    sheet1["A1"] = "기사 주제"
    sheet1["B1"] = "기사 제목"
    sheet1["C1"] = "기사 링크"
    sheet1["D1"] = "기사 내용"
    sheet1["E1"] = "좋아요"
    sheet1["F1"] = "훈훈해요"
    sheet1["G1"] = "슬퍼요"
    sheet1["H1"] = "화나요"
    sheet1["I1"] = "기대 돼요"
    sheet1["J1"] = "트래픽 양"

    wb.save(new_file_name)
    print("엑셀파일 생성")

# 엑셀 파일 쓰기
def write_xls(sheet_title, file_name, tab_name, list):
    load_wb = openpyxl.load_workbook(file_name, data_only=True)
    # 시트 이름으로 불러오기
    load_sht = load_wb[sheet_title]
    print("---엑셀화 중...")
    feel_list = ["E","F","G","H","I"]
    for i in range(len(list)):
        #print("현재 엑셀 rows 현황 : " + str(load_sht.max_row))
        now_sheet_row = load_sht.max_row
        load_sht["A" + str(now_sheet_row+1)] = tab_name
        load_sht["B" + str(now_sheet_row+1)] = list[i]['title']
        load_sht["C" + str(now_sheet_row+1)] = list[i]['current_url']
        load_sht["D" + str(now_sheet_row+1)] = list[i]['e_content']

        for j in range(5):
            if list[i]['feel'][j] == "":
                list[i]['feel'][j] = 0
            else:
                load_sht[str(feel_list[j]) + str(now_sheet_row+1)] = list[i]['feel'][j]
                # load_sht["F" + str(now_sheet_row+1)] = list[i]['feel'][1]
                # load_sht["G" + str(now_sheet_row+1)] = list[i]['feel'][2]
                # load_sht["H" + str(now_sheet_row+1)] = list[i]['feel'][3]
                # load_sht["I" + str(now_sheet_row+1)] = list[i]['feel'][4]

    load_wb.save(file_name)
    print("##내용 저장 완료")


#===================================================================================================================================================================
#현재 선택한 좌측 메뉴의 끝 페이지까지 이동한다.
def move_end_content(driver, social_tab, social_tab_under):
    #driver = setting_driver()
    time.sleep(2)
    print("---Open Page")
    driver.get("https://news.naver.com/")
    time.sleep(2)
    print(social_tab)
    print(social_tab_under)

    driver.find_element(By.XPATH,"/html/body/section/header/div[2]/div/div/div[1]/div/div/ul/li[4]").click()
    time.sleep(2)
    #메인 페이지 상위 메뉴중 사회메뉴를 클릭한다.
    social_tab_list = driver.find_element(By.XPATH, "//*[@id='lnb']/ul")
    tab_child = social_tab_list.find_elements(By.XPATH, ".//*")

    for i in range(len(tab_child)):
        if tab_child[i].text == social_tab:
            print("---상위 메뉴 이동 : " + tab_child[i].text)
            tab_child[i].click()
            break
    # driver.find_element(By.XPATH, "/html/body/section/header/div[2]/div/div/div[1]/div/div/ul/li[4]/a/span").click()
    time.sleep(2)

    social_tab_under_list = driver.find_element(By.XPATH, "//*[@id='snb']/ul")
    tab_child = social_tab_under_list.find_elements(By.XPATH, ".//*")
    for i in range(len(tab_child)):
        if tab_child[i].text == social_tab_under:
            print("----좌측 메뉴 이동 : " + tab_child[i].text)
            tab_child[i].click()
            break
    #driver.find_element(By.XPATH, "//*[@id='snb']/ul/li[10]/a").click()
    time.sleep(2)

    #테스트 코드 2021-12-20
    driver.execute_script("window.scrollTo(0, (document.body.scrollHeight));")
    area_pg_btn = driver.find_element(By.XPATH, "//*[@id='main_content']/div[4]")
    day_child = area_pg_btn.find_elements(By.XPATH, ".//*")
    print(day_child[2].text)
    day_child[2].click()
    time.sleep(2)

    while(1):
        test_dp = ""

        driver.execute_script("window.scrollTo(0, (document.body.scrollHeight));")
        #페이지 버튼 클래스 네임에 특수문자가 있어서 추적이 안된다.
        #페이지 버튼을 포함하는 객체의 XPath를 이용해 자식요소인 페이지 버튼들을 추적함
        area_pg_btn = driver.find_element(By.XPATH,"// *[ @ id = 'main_content'] / div[3]")
        btn_child = area_pg_btn.find_elements(By.XPATH, ".//*")
        #print(btn_child)
        print("현재 페이지 수 : " + str(len(btn_child)))
        # for i in range(len(btn_child)):
        #     print(str(i) + "번째 : " + btn_child[i].text)

        if len(btn_child) < 11 and btn_child[len(btn_child)-1].text != "다음":
            try:
                print("마지막 페이지 : " + str(btn_child[len(btn_child)-1].text))
                btn_child[len(btn_child)-1].click()
                print("마지막 페이지입니다.")
                break
            except:
                break

        if len(btn_child) < 12 and btn_child[0].text == "이전":
            try:
                print("마지막 페이지 : " + str(btn_child[len(btn_child)-1].text))
                btn_child[len(btn_child)-1].click()
                print("마지막 페이지입니다.")
                break
            except:
                break

        # 임시 테스트용 2021-12-20
        # for i in range(len(btn_child)):
        #     if btn_child[i].text == "49":
        #         print("임시테스트 페이지로 이동합니다.")
        #         btn_child[i].click()
        #         test_dp = "123"
        #         break
        # if test_dp == "123":
        #     break

        for i in range(len(btn_child)):
            #print(btn_child[i].text)
            if btn_child[i].text == "다음":
                btn_child[i].click()
                print("다음페이지로 이동합니다.")

        time.sleep(3)



#===================================================================================================================================================================
#현재 선택한 좌측 메뉴의 끝 페이지까지 이동한다.
def move_prve_content(driver, sheet_title, file_name, social_tab_under):
    news_list = []
    #여기에 Crop Content가 나올것임
    time.sleep(2)

    while(1):
        time.sleep(2)
        driver.execute_script("window.scrollTo(0, (document.body.scrollHeight));")
        time.sleep(2)
        #페이지 버튼 클래스 네임에 특수문자가 있어서 추적이 안된다.
        #페이지 버튼을 포함하는 객체의 XPath를 이용해 자식요소인 페이지 버튼들을 추적함
        area_pg_btn = driver.find_element(By.XPATH, "// *[ @ id = 'main_content'] / div[3]")
        btn_child = area_pg_btn.find_elements(By.XPATH, ".//*")
        now_page = driver.current_url.split("=")[-1]
        print("=== [" + str(social_tab_under) + "]" + "현재 페이지 : " + str(now_page) + " PAGE ===")
        #현재 페이지가 1이라면 페이지 이동 자체를 멈춘다.
        #page_dictionary = crop_content(driver, social_tab_under)
        #write_xls(sheet_title, file_name, social_tab_under, page_dictionary)
        # if str(now_page) == "1":
        #     print("드라이버 종료")
        #     driver.quit()
        #     break
        
        for b in range(len(news_list)):
            print(news_list[b])

        if str(now_page) == "1":
            print("마지막 페이지 입니다.")
            break

        #현재 페이지-1 페이지로 이동한다.
        for a in range(len(btn_child)):
            print("[" + str(social_tab_under) + "]" + "[" + str(a) + "] : " + btn_child[a].text)
        for i in range(len(btn_child)):
            #print("페이지 검사중 : " + btn_child[i].text)
            if btn_child[i].text == str(int(now_page)-1):
                print(btn_child[i].text + "페이지로 이동")
                btn_child[i].send_keys(Keys.ENTER)
                break
            #현재 페이지-1 페이지로 이동하다가 더이상 갈수있는 페이지가 없을때 이전버튼을 선택한다.
            
            if btn_child[1].text == str(now_page) and btn_child[0].text =="이전":
                print( str(social_tab_under) + " 페이지 한단락을 이동합니다")
                btn_child[0].send_keys(Keys.ENTER)
                break

        time.sleep(3)



#현재 활성화된 페이지의 뉴스 리스트 정보를 반환한다.
#한번 실행시 마다 뉴스 한개의 정보를 담은 딕셔너리를 반환함
def crop_content(driver, social_tab_under):
    #driver = setting_driver()
    new_list = []
    news_dictionary = {}
    print("---Crop Start")
    #driver.get("https://news.naver.com/main/list.naver?mode=LS2D&mid=shm&sid2=249&sid1=102&date=20211218&page=3")
    time.sleep(2)
    #driver.execute_script("window.scrollTo(0, (document.body.scrollHeight/2));")
    news_list = driver.find_elements(By.CLASS_NAME,"photo")
    print("[" + str(social_tab_under) + "]" + "뉴스 리스트 출력 : " + str(len(news_list)))
    time.sleep(2)

    for i in range(len(news_list)):
        ActionChains(driver).key_down(Keys.CONTROL) \
            .click(news_list[i]) \
            .key_up(Keys.CONTROL) \
            .perform()

        driver.switch_to.window(driver.window_handles[-1])
        time.sleep(1)
        news_dictionary = {}
        #print("=================================  " + str(i) + "  =========================================")
        try:
            news_title = driver.find_element(By.CLASS_NAME,"tts_head").text
            #print("----뉴스 제목 : " + news_title)
            # print("----뉴스 링크 : " + driver.current_url)
            e_content = driver.find_element(By.XPATH, "//*[@id='articleBodyContents']").text
            # print("----뉴스 내용" + e_content)

            feel_list = driver.find_elements(By.CLASS_NAME, "u_likeit_list_count")
            # print("감정표현 출력 수 : " + str(len(feel_list)))

            # 출력 결과가 10개 중 5~10 값이 실제 감정 표현 값이다.
            emotion_list = ["좋아요", "훈훈해요", "슬퍼요", "화나요", "후속기사 원해요"]
            emotion_list_value = []

            #혹시 들고온 감정 리스트가 10개가 안된다 == 로딩에러가 났다면 일관적으로 모두 0으로 받음
            if len(feel_list) < 10:
                emotion_list_value = [0, 0, 0, 0, 0]
            else:
                for j in range(5, len(feel_list)):
                    # print(emotion_list[j-5] + " : " + feel_list[j].text)
                    emotion_list_value.append(feel_list[j].text.replace(",", ""))
            
            #특정 정치뉴스는 좋아요가 추적되지 않아 다음과 같이 대체

            if not emotion_list_value:
                emotion_list_value = [0, 0, 0, 0, 0]

            news_dictionary["title"] = news_title
            news_dictionary["current_url"] = driver.current_url
            news_dictionary["e_content"] = e_content
            news_dictionary["feel"] = emotion_list_value
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
            time.sleep(2)

            new_list.append(news_dictionary)
        except:
            print("다른유형의 기사입니다. 해당 기사는 오류로 인해 건너뜁니다.")
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
            time.sleep(2)


        #print(news_dictionary)

        # for a in range(len(new_list)):
        #     print(new_list[a])
        
    return new_list

def send_mail(from_mail, to_mail, attach):
    # 세션생성, 로그인
    print("메일 발송중 1")
    s = smtplib.SMTP('smtp.gmail.com', 587, None, 30)
    s.starttls()
    s.login('today22motion@gmail.com', 'rnjrnjwm12')
    print("메일 발송중 2")
    # 제목, 본문 작성
    msg = MIMEMultipart()
    today = datetime.today().strftime("%Y-%m-%d")
    msg['Subject'] = Header(s=(date.today() - timedelta(1)).isoformat() + '뉴스 분석데이터 발송.', charset='utf-8')
    msg['From'] = from_mail
    msg['To'] = to_mail
    body = MIMEText((date.today() - timedelta(1)).isoformat() + '뉴스 분석데이터 발송합니다. 문제가 있을경우 개인연락 부탁드립니다.', _charset='utf-8')
    msg.attach(body)
    attach_file = attach
    print("메일 발송중 3")
    part = MIMEBase('application', "octet-stream")
    part.set_payload(open(attach_file,"rb").read())
    encode_base64(part)
    part.add_header('Content-Disposition', 'attachment; filename="%s"' % os.path.basename(attach_file))
    msg.attach(part)
    print("메일 발송중 4")
    # 메일 전송
    s.send_message(msg)
    s.quit()
    print("---send email with attach_file 'To. '" + to_mail + "'")



def crolling_start(sheet_title, file_name, social_tab, social_tab_under):
    print("##Setting Display")
    #display = Display(visible=0, size=(1920, 1000))
    #display.start()
    driver = setting_driver()
    #특정 뉴스 탭에 들어가 맨 끝페이지 까지 이동함
    move_end_content(driver, social_tab, social_tab_under)
    #특정 뉴스 탭의 마지막에 도착하면 앞으로 한 페이지씩 이동하며 뉴스 정보를 스크랩함
    print("첫번째 페이지까지 자동 스크랩을 실행합니다.")
    move_prve_content(driver, sheet_title, file_name, social_tab_under)
    #print(list)
    #display.stop()
    driver.quit()
    print("드라이버를 종료합니다.")
    #print("######################## 리스트 내역 엑셀화 ##############################")

    




