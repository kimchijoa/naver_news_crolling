import pymysql as mysql
import openpyxl
import time

def create_conn():
    conn = mysql.connect(
        user='kkh', 
        passwd='rnjrnjwm12', 
        host='172.19.0.3', 
        db='today_emotion', 
        charset='utf8',
        port=3306
    )

    sql_cursor = conn.cursor()
    print("DB 연결")
    return sql_cursor, conn

#===============================================================================================================================
#오늘 수집한 엑셀데이터 DB에 입력함
def insert_total_data(conn, sql_cursor, sheet_title, file_name, y_date):
    sql = "INSERT INTO today_news_data (n_category, n_title, n_link, n_content, n_e_like, n_e_good, n_e_sad, n_e_angry, n_e_expect, news_date) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
    try:
        load_wb = openpyxl.load_workbook(file_name, data_only=True)
    except:
        print(str(file_name) + " 파일이 존재하지 않습니다. DB입력을 종료합니다.")
        return 0

    # 시트 이름으로 불러오기
    load_sht = load_wb[sheet_title]
    #최대 row 구하기
    now_sheet_row = load_sht.max_row
    #최대 row만큼 insert 반복하기
    fail_count = 0
    for i in range(2, now_sheet_row+1):
        category_name = load_sht["A" + str(i)].value
        news_title = load_sht["B" + str(i)].value
        news_link = load_sht["C" + str(i)].value
        news_content = load_sht["D" + str(i)].value
        n_e_like = load_sht["E" + str(i)].value
        n_e_good = load_sht["F" + str(i)].value
        n_e_sad = load_sht["G" + str(i)].value
        n_e_angry = load_sht["H" + str(i)].value
        n_e_expect = load_sht["I" + str(i)].value
        news_date = y_date
        val = (category_name, news_title, news_link, news_content, n_e_like, n_e_good, n_e_sad, n_e_angry, n_e_expect, news_date)
        #time.sleep(1)
        #print(str(val))
        try:
            sql_cursor.execute(sql, val)
            conn.commit()
            print("데이터 삽입 [" + str(i) + "]")
        except:
            print("실패")
            fail_count=+1
        

    conn.close()
    print("삽입 실패한 데이터 갯수 : " + str(fail_count) + "개")

#===============================================================================================================================
#오늘 수집한 엑셀데이터 DB에 입력함
def insert_aws_data(conn, sql_cursor, s3_key, s3_file_name, s3_file_size, update_date):
    sql = "INSERT INTO s3_info (s3_key, s3_file_name, s3_file_size, update_date) VALUES (%s, %s, %s, %s)"
    val = (s3_key, s3_file_name, s3_file_size, update_date)
    try:
        sql_cursor.execute(sql, val)
        conn.commit()
        print("s3 데이터가 성공적으로 업로드 되었습니다.")
    except:
        print("s3 업로드 데이터 실패")
    
    conn.close()


#===================================================================================================================================

#엑셀 파일을 순차적으로 DB에 저장함
def read_grap_speed(sheet_title, file_name):
    # sheet_title = "Sheet1"
    # file_name = "total_news_data_2021-12-28.xlsx"
    try:
        load_wb = openpyxl.load_workbook(file_name, data_only=True)
         # 시트 이름으로 불러오기
        load_sht = load_wb[sheet_title]
        #최대 row 구하기
        now_sheet_row = load_sht.max_row
        #최대 row만큼 insert 반복하기
        fail_count = 0
        for i in range(2, now_sheet_row+1):
            category_name = load_sht["A" + str(i)].value
            news_title = load_sht["B" + str(i)].value
            news_link = load_sht["C" + str(i)].value
            news_content = load_sht["D" + str(i)].value
            n_e_like = load_sht["E" + str(i)].value
            n_e_good = load_sht["F" + str(i)].value
            n_e_sad = load_sht["G" + str(i)].value
            n_e_angry = load_sht["H" + str(i)].value
            n_e_expect = load_sht["I" + str(i)].value
            news_date = "2021-12-28"
            val = (category_name, news_title, news_link, news_content, n_e_like, n_e_good, n_e_sad, n_e_angry, n_e_expect, news_date)
            try:
                sql_cursor.execute(sql, val)
                conn.commit()
                print("데이터 삽입 [" + str(i) + "]")
            except:
                print("실패")
                fail_count=+1
            

        conn.close()
        print("삽입 실패한 데이터 갯수 : " + str(fail_count) + "개")
    except:
        print(str(file_name) + " 파일이 존재하지 않습니다.")
   



# create_conn()
