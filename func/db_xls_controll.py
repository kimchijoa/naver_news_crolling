import collections
from sys import hash_info
import time
from typing import Collection
from selenium import webdriver
#from pyvirtualdisplay import Display
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
import chromedriver_autoinstaller
from selenium.webdriver.common.keys import Keys
import openpyxl
import mysql_conn as mysql
import pymysql
from datetime import datetime
from datetime import date, timedelta
#===================== 메일 발송 패키지
import os
import smtplib
from email.encoders import encode_base64
from email.header import Header
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import shutil
import json
from collections import OrderedDict
#===============================================================================================================================
def create_xls(sheet_title, file_name):
    wb = openpyxl.Workbook()
    sheet1 = wb.active
    sheet1.title = sheet_title
    new_file_name = file_name

    sheet1["A1"] = "기사 주제"
    sheet1["B1"] = "기사 제목"
    sheet1["C1"] = "기사 링크"
    sheet1["D1"] = "기사 내용"
    sheet1["E1"] = "좋아요"
    sheet1["F1"] = "훈훈해요"
    sheet1["G1"] = "슬퍼요"
    sheet1["H1"] = "화나요"
    sheet1["I1"] = "기대 돼요"
    sheet1["J1"] = "날짜"

    wb.save(new_file_name)
    print("엑셀파일 생성")

#엑셀 파일 쓰기 ====================================================================================================
def write_xls(sheet_title, file_name, tab_name, value_list):
    load_wb = openpyxl.load_workbook(file_name, data_only=True)
    # 시트 이름으로 불러오기
    load_sht = load_wb[sheet_title]
    feel_list = ["E","F","G","H","I"]
    for i in range(len(list)):
        #print("현재 엑셀 rows 현황 : " + str(load_sht.max_row))
        now_sheet_row = load_sht.max_row
        load_sht["A" + str(now_sheet_row+1)] = tab_name
        load_sht["B" + str(now_sheet_row+1)] = value_list[i]['title']
        load_sht["C" + str(now_sheet_row+1)] = value_list[i]['current_url']
        load_sht["D" + str(now_sheet_row+1)] = value_list[i]['e_content']

        for j in range(5):
            if value_list[i]['feel'][j] == "":
                value_list[i]['feel'][j] = 0
            else:
                load_sht[str(feel_list[j]) + str(now_sheet_row+1)] = value_list[i]['feel'][j]

    load_wb.save(file_name)


create_xls("Sheet1", "news_data/2022-01-14_total_news_data_db_test.xlsx")
file_name = "news_data/2022-01-14_total_news_data_db_test.xlsx"
#DB 연결 설정
sql_cursor, conn = mysql.create_conn()
sql = "SELECT * FROM today_news_data"
cursor = conn.cursor(pymysql.cursors.DictCursor)
cursor.execute(query=sql)
print("DB내역 가져오는 중...")
result = cursor.fetchall()
#print(result["idx"])
load_wb = openpyxl.load_workbook(file_name, data_only=True)
load_sht = load_wb["Sheet1"]

for i in range(len(result)):
    #row가 딕셔너리 형태로 out 됨
    print(str(result[i]["idx"]) + " / Title : " + result[i]["n_title"])
    #print("현재 엑셀 rows 현황 : " + str(load_sht.max_row))
    #n_category, n_title, n_link, n_content, n_e_like, n_e_good, n_e_sad, n_e_angry, n_e_expect, news_date
    now_sheet_row = load_sht.max_row
    load_sht["A" + str(now_sheet_row+1)] = result[i]["n_category"]
    load_sht["B" + str(now_sheet_row+1)] = result[i]["n_title"]
    load_sht["C" + str(now_sheet_row+1)] = result[i]["n_link"]
    load_sht["D" + str(now_sheet_row+1)] = result[i]["n_content"]
    load_sht["E" + str(now_sheet_row+1)] = result[i]["n_e_like"]
    load_sht["F" + str(now_sheet_row+1)] = result[i]["n_e_good"]
    load_sht["G" + str(now_sheet_row+1)] = result[i]["n_e_sad"]
    load_sht["H" + str(now_sheet_row+1)] = result[i]["n_e_angry"]
    load_sht["I" + str(now_sheet_row+1)] = result[i]["n_e_expect"]
    load_sht["j" + str(now_sheet_row+1)] = result[i]["news_date"]
    #time.sleep(0.5)
load_wb.save(file_name)
print("총합 db 엑셀 파일 생성완료")
