import collections
from sys import hash_info
import time
from typing import Collection
from selenium import webdriver
#from pyvirtualdisplay import Display
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
import chromedriver_autoinstaller
from selenium.webdriver.common.keys import Keys
import openpyxl
import mysql_conn as mysql
import pymysql
from datetime import datetime
from datetime import date, timedelta
#===================== 메일 발송 패키지
import os
import smtplib
from email.encoders import encode_base64
from email.header import Header
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import shutil
import json
from collections import OrderedDict
#===============================================================================================================================
def create_xls(sheet_title, file_name):
    wb = openpyxl.Workbook()
    sheet1 = wb.active
    sheet1.title = sheet_title
    new_file_name = file_name

    sheet1["A1"] = "기사 주제"
    sheet1["B1"] = "기사 제목"
    sheet1["C1"] = "기사 링크"
    sheet1["D1"] = "기사 내용"
    sheet1["E1"] = "좋아요"
    sheet1["F1"] = "훈훈해요"
    sheet1["G1"] = "슬퍼요"
    sheet1["H1"] = "화나요"
    sheet1["I1"] = "기대 돼요"
    sheet1["J1"] = "날짜"

    wb.save(new_file_name)
    print("엑셀파일 생성")


def create_folder(root_folder_name, folder_name, sheet_title, file_name_list):

    try:
        if not os.path.exists(root_folder_name):
            os.makedirs(root_folder_name)
    except OSError:
        print("root 폴더를 생성할수 없습니다.")

    try:
        if not os.path.exists(folder_name):
            #shutil.rmtree(folder_name)
            os.makedirs(folder_name)
            for i in range(len(file_name_list)):
                create_xls(sheet_title, file_name_list[i])
            print("폴더가 생성되었습니다.")

        elif os.path.exists(folder_name):
            print("기존 폴더를 삭제하고 재생성 합니다.")
            shutil.rmtree(folder_name)
            os.makedirs(folder_name)
            for i in range(len(file_name_list)):
                create_xls(sheet_title, file_name_list[i])
            print("폴더가 생성되었습니다.")

    except OSError:
        print("Error: 폴더를 생성할수 없습니다.")


#엑셀 파일 쓰기 ====================================================================================================
def write_xls(sheet_title, file_name, tab_name, value_list):
    load_wb = openpyxl.load_workbook(file_name, data_only=True)
    # 시트 이름으로 불러오기
    load_sht = load_wb[sheet_title]
    feel_list = ["E","F","G","H","I"]
    for i in range(len(value_list)):
        #print("현재 엑셀 rows 현황 : " + str(load_sht.max_row))
        now_sheet_row = load_sht.max_row
        load_sht["A" + str(now_sheet_row+1)] = tab_name
        load_sht["B" + str(now_sheet_row+1)] = value_list[i]['title']
        load_sht["C" + str(now_sheet_row+1)] = value_list[i]['current_url']
        load_sht["D" + str(now_sheet_row+1)] = value_list[i]['e_content']

        for j in range(5):
            if value_list[i]['feel'][j] == "":
                value_list[i]['feel'][j] = 0
            else:
                load_sht[str(feel_list[j]) + str(now_sheet_row+1)] = value_list[i]['feel'][j]

    load_wb.save(file_name)

#=======================================================================================
#그래프 속도 측정용 엑셀 파일 함수
def create_graph_info_xls(sheet_title, file_name):
    wb = openpyxl.Workbook()
    sheet1 = wb.active
    sheet1.title = sheet_title
    new_file_name = file_name

    sheet1["A1"] = "Category"
    sheet1["B1"] = "list_count"
    sheet1["C1"] = "cost_time"
    sheet1["D1"] = "update_date"

    wb.save(new_file_name)
    print("엑셀파일 생성")

#그래프 속도 측정용 엑셀파일 쓰기 ====================================================================================================
def write_graph_info_xls(sheet_title, file_name, tab_name, list_count, cost_value, update_date):
    load_wb = openpyxl.load_workbook(file_name, data_only=True)
    # 시트 이름으로 불러오기
    load_sht = load_wb[sheet_title]

    now_sheet_row = load_sht.max_row
    load_sht["A" + str(now_sheet_row+1)] = tab_name
    load_sht["B" + str(now_sheet_row+1)] = list_count
    load_sht["C" + str(now_sheet_row+1)] = cost_value
    load_sht["D" + str(now_sheet_row+1)] = update_date
    load_wb.save(file_name)


def load_local_c_speed_info(date):
    file_name = "/naver_news_crolling/news_data/news_data_" + date + "/" + "graph_speed_info_" + date  + ".xlsx"
    sheet_title = "today_crolling_speed"
    load_wb = openpyxl.load_workbook(file_name, data_only=True)
    load_sht = load_wb[sheet_title]
    
    return load_sht.max_row - 1
    

def load_local_total_news_info(date):
    file_name = "/naver_news_crolling/news_data/news_data_" + date + "/" + "total_news_data_" + date  + ".xlsx"
    sheet_title = "Sheet1"
    load_wb = openpyxl.load_workbook(file_name, data_only=True)
    load_sht = load_wb[sheet_title]
    
    return load_sht.max_row - 1
    